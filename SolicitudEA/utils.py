import openpyxl
from openpyxl.utils import get_column_letter
from config import CAMPOS, HOJA_SOLICITUD

def cargar_plantilla(ruta_plantilla):
    """Carga el archivo plantilla de solicitud"""
    return openpyxl.load_workbook(ruta_plantilla)

def cargar_datos(ruta_datos):
    """Carga el archivo de datos fuente"""
    return openpyxl.load_workbook(ruta_datos)

def buscar_y_reemplazar(hoja, texto_buscar, texto_reemplazar):
    """Busca y reemplaza texto en una hoja de Excel"""
    for row in hoja.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and texto_buscar in cell.value:
                cell.value = cell.value.replace(texto_buscar, texto_reemplazar)

def procesar_solicitud(wb_plantilla, datos_fila):
    """Procesa la plantilla con los datos de una fila"""
    hoja = wb_plantilla[HOJA_SOLICITUD]
    
    # Reemplazar todos los campos marcados
    for campo_db, campo_plantilla in CAMPOS.items():
        valor = datos_fila.get(campo_db, '')
        buscar_y_reemplazar(hoja, campo_plantilla, str(valor))
    
    return wb_plantilla

def guardar_solicitud(wb, nombre_archivo, ruta_output):
    """Guarda la solicitud generada"""
    ruta_completa = f"{ruta_output}/{nombre_archivo}"
    wb.save(ruta_completa)
    return ruta_completa